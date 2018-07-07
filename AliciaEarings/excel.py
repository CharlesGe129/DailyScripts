import openpyxl
import openpyxl.styles as sty
import os
import time
import datetime
import sys


class Excel:
    def __init__(self, path='', today=''):
        if today == '':
            os.environ['TZ'] = 'Asia/Shanghai'
            time.tzset()
            self.sheet_date = time.strftime('%m.%d')
        else:
            self.sheet_date = today
        self.sheet_yesterday = self.cal_yesterday(self.sheet_date)
        print("Sheet名字: 今天是{0}, 昨天是{1}".format(self.sheet_date, self.sheet_yesterday))

        self.path = path
        f_order = openpyxl.load_workbook(path + '/order.xlsx')
        self.f_order = f_order
        self.sheet_order = f_order[self.sheet_date] if self.sheet_date in f_order.sheetnames \
            else f_order.create_sheet(self.sheet_date)
        f_shipping = openpyxl.load_workbook(path + '/shipping.xlsx')
        self.f_shipping = f_shipping
        self.sheet_shipping = f_shipping[self.sheet_date] if self.sheet_date in f_shipping.sheetnames \
            else f_shipping.create_sheet(self.sheet_date)
        f_product = openpyxl.load_workbook(path + '/product.xlsx')
        self.f_product = f_product
        self.sheet_product = f_product[self.sheet_date] if self.sheet_date in f_product.sheetnames \
            else f_product.create_sheet(self.sheet_date)
        self.sheet_storage = f_product['产品总表']
        f_backup = openpyxl.load_workbook(path + '/backup.xlsx')
        self.f_backup = f_backup
        self.sheet_backup = f_backup[self.sheet_yesterday]
        self.sheet_backup_today = f_backup[self.sheet_date] if self.sheet_date in f_backup.sheetnames \
            else f_backup.create_sheet(self.sheet_date)
        self.data_storage = dict()
        self.parse_storage()

    def parse_storage(self):
        table = self.sheet_backup
        indices = self.get_indices(table)
        for row in range(2, table.max_row + 1):
            id = table.cell(row, indices['序号']).value
            self.data_storage[id] = dict()
            self.data_storage[id]['name'] = table.cell(row, indices['名字']).value
            cost = table.cell(row, indices['成本']).value
            self.data_storage[id]['cost'] = float(cost) if cost else 0.0
            price = table.cell(row, indices['单价']).value
            self.data_storage[id]['price'] = float(price) if price else 0.0
            self.data_storage[id]['profit'] = self.data_storage[id]['price'] - self.data_storage[id]['cost']
            storage = table.cell(row, indices['库存']).value
            self.data_storage[id]['storage'] = storage if storage else 0

    @staticmethod
    def cal_yesterday(today):
        year = int(time.strftime('%Y'))
        month = int(today.split('.')[0])
        day = int(today.split('.')[1])
        previous = datetime.datetime(year, month, day) + datetime.timedelta(days=-1)
        return previous.strftime('%m.%d')

    def cal_today_storage(self):

        def update_storage():
            table = self.sheet_storage
            indices = self.get_indices(table)
            for row in range(2, table.max_row + 1):
                table.cell(row, col_storage).value = self.data_storage[table.cell(row, indices['序号']).value]['storage']
            for row in range(2, self.sheet_storage.max_row + 1):
                num_storage = self.sheet_storage.cell(row, col_storage).value
                num_storage = 0 if not num_storage else int(num_storage)
                if num_storage <= 0:
                    for col in range(1, self.sheet_storage.max_column + 1):
                        self.sheet_storage.cell(row, col).fill = sty.PatternFill(patternType='solid', fgColor="fa8072")
                else:
                    for col in range(1, self.sheet_storage.max_column + 1):
                        self.sheet_storage.cell(row, col).fill = sty.PatternFill(fgColor="ffffff")
            self.f_product.save(self.path + '/product.xlsx')
            print("已更新今日最新库存")

        def update_backup():
            table_backup_today = self.f_backup[self.sheet_date]
            self.copy_from_sheet_to_new(self.sheet_storage, table_backup_today, self.f_backup,
                                        '/backup.xlsx')
            for row in range(2, table_backup_today.max_row + 1):
                num_storage = table_backup_today.cell(row, col_storage).value
                num_storage = 0 if not num_storage else int(num_storage)
                if num_storage <= 0:
                    for col in range(1, table_backup_today.max_column + 1):
                        table_backup_today.cell(row, col).fill = sty.PatternFill(patternType='solid', fgColor="fa8072")
            self.f_backup.save(self.path + '/backup.xlsx')
            print("已备份今日最新库存")

        self.copy_from_sheet_to_new(self.sheet_backup, self.sheet_storage, self.f_product, '/product.xlsx')
        print("已拉取昨日{0}的备份库存作为今日的计算原数据".format(self.sheet_yesterday))
        table = self.sheet_product
        indices = self.get_indices(table)
        col_sell = indices['售出件数'] if '售出件数' in indices else 0
        col_id = indices['产品编号']
        # data_storage = [id, storage]
        col_storage = self.get_indices(self.sheet_backup)['库存']
        data_sell = list()
        # Prepare selling information: [id starts from 1, sum of sell, if sell or buy]
        for i in range(2, table.max_row + 1):
            if table.cell(i, col_id).value is None:
                print('"Product.xlsx"的"{0}"表, 第{1}行缺少产品id，跳过该行'.format(self.sheet_date, i))
                continue
            data_sell.append([table.cell(i, col_id).value, int(table.cell(i, col_sell).value),
                              table.cell(i, indices['微信名']).value != '进货'])
        # Calculate today's storage
        for entry in data_sell:
            self.data_storage[entry[0]]['storage'] -= entry[1] if entry[2] else -entry[1]
        # Write back to xlsx
        update_storage()
        update_backup()
        return {'success': True}

    def cal_today_selling(self):
        table = self.sheet_product
        indices = self.get_indices(table)
        for i in range(2, table.max_row + 1):
            if table.cell(i, indices['微信名']).value == '进货':
                continue
            if not self.check_required_fields(table, i):
                continue
            num = int(table.cell(i, indices['售出件数']).value)

            id = table.cell(i, indices['产品编号']).value
            price = self.data_storage[id]['price']
            cost = self.data_storage[id]['cost']
            profit = self.data_storage[id]['profit']
            bonus = float(table.cell(i, indices['优惠']).value)

            total = num * price
            total_cost = num * cost
            original_profit = total - total_cost
            total_final = total * bonus
            total_profit = total_final - total_cost
            table.cell(i, indices['单价']).value = price
            table.cell(i, indices['成本']).value = cost
            table.cell(i, indices['利润']).value = profit
            table.cell(i, indices['总金额']).value = total
            table.cell(i, indices['总利润']).value = original_profit
            table.cell(i, indices['折后金额']).value = total_final
            table.cell(i, indices['折后利润']).value = total_profit
        self.f_product.save(self.path + '/product.xlsx')

    def cal_order_group_by_buyer(self):
        table = self.sheet_product
        indices = self.get_indices(table)
        data = dict()
        for i in range(2, table.max_row + 1):
            if table.cell(i, indices['微信名']).value == '进货':
                continue
            id = table.cell(i, indices['微信名']).value
            entry = [table.cell(i, j).value if table.cell(i, j).value else '' for j in range(1, table.max_column + 1)]
            if not self.check_required_fields(table, i):
                # Copy this line
                if id in data:
                    data[id]['entry'].append(entry)
                else:
                    data[id] = dict()
                    data[id]['sum'] = 0
                    data[id]['entry'] = [entry]
                continue
            total = table.cell(i, indices['折后金额']).value
            if id in data:
                data[id]['sum'] += total
                data[id]['entry'].append(entry)
            else:
                data[id] = dict()
                data[id]['sum'] = total
                data[id]['entry'] = [entry]
        i = 2
        for id, value in data.items():
            for each_row in value['entry']:
                for j in range(len(each_row)):
                    table.cell(i, j + 1).value = each_row[j]
                table.cell(i, indices['折后利润'] + 1).value = value['sum']
                i += 1
        if table.max_row > 1:
            table.cell(1, indices['折后利润'] + 1).value = '该用户今日总购买金额（若一日多单请老婆大人自行计算~）'
        self.f_product.save(self.path + '/product.xlsx')
        print("已初步整理今日每个微信买家的产品信息及总购买金额。若一日多单请老婆大人自行计算该单需要额外支付多少钱~")

    def check_required_fields(self, table, i):
        indices = self.get_indices(table)
        success = True
        if not table.cell(i, indices['售出件数']).value:
            success = False
            print('"Product.xlsx"的"{0}"表, 第{1}行缺少"售出件数"，跳过该行'.format(self.sheet_date, i))
        if not table.cell(i, indices['优惠']).value:
            success = False
            print('"Product.xlsx"的"{0}"表, 第{1}行缺少"优惠"，跳过该行'.format(self.sheet_date, i))
        return success

    def copy_from_sheet_to_new(self, sheet_ori, sheet_des, file, filename):
        for i in range(1, sheet_des.max_row + 1):
            for j in range(1, sheet_des.max_column + 1):
                sheet_des.cell(i, j).value = ''
        for i in range(1, sheet_ori.max_row + 1):
            for j in range(1, sheet_ori.max_column + 1):
                sheet_des.cell(i, j).value = sheet_ori.cell(i, j).value
        file.save(self.path + filename)

    @staticmethod
    def get_indices(table):
        indices = dict()
        for i in range(1, table.max_column + 1):
            indices[table.cell(1, i).value] = i
        return indices

    def all_storage(self):
        table = self.sheet_storage
        indices = self.get_indices(table)
        col_index = indices['序号']
        col_name = indices['名字']
        col_num = indices['库存']
        for i in range(2, table.max_row + 1):
            print("#{0}, {1}, 库存: {2}".format(table.cell(i, col_index).value, table.cell(i, col_name).value,
                                              table.cell(i, col_num).value))


path = '/Users/charlesge/Downloads/AliciaEarings'

e = Excel(path=path, today=sys.argv[1] if len(sys.argv) == 2 else '')
response = e.cal_today_storage()
if not response['success']:
    print(response['msg'])
e.cal_today_selling()
e.cal_order_group_by_buyer()
e.all_storage()
